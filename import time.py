import os
import re
import time
import json
import shutil
import tempfile
import subprocess
import unicodedata
from urllib.parse import urlparse, urljoin

import openpyxl
import requests
from bs4 import BeautifulSoup, FeatureNotFound

API_KEY = os.getenv("GOOGLE_MAPS_API_KEY", "SUA_KEY_AQUI")
ARQ_IN = r"Prospecção Novos Clientes.xlsx"
ARQ_OUT = r"Prospecção Novos Clientes - preenchido.xlsx"

SHEET = "Clientes"
HEADER_ROW = 1
TOP_N = 12
SLEEP = 0.25
MAX_ROWS = int(os.getenv("MAX_ROWS", "0"))  # 0 = processa tudo

PHONE_RE = re.compile(
    r"(?:(?:\+?55)\s*)?"
    r"(?:\(?\d{2}\)?\s*)?"
    r"(?:9?\d{4})[-\s]?\d{4}",
    re.IGNORECASE,
)

ADDR_HINT_RE = re.compile(
    r"(Rua|Avenida|Av\.|Rodovia|R\.|Alameda|Travessa|Estrada|Praça|Quadra|Lote|Km)\b.*?\d{1,5}",
    re.IGNORECASE,
)

COMPANY_SUFFIX_RE = re.compile(
    r"\b(SA|S\.?A\.?|LTDA|EIRELI|ME|EPP|INDUSTRIA|INDUSTRIAL|COMERCIO|COMERCIAL|EMBALAGENS?)\b",
    re.IGNORECASE,
)

SESSION = requests.Session()
SESSION.headers.update(
    {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36"
    }
)


def has_api_key() -> bool:
    return bool(API_KEY and API_KEY != "SUA_KEY_AQUI")


def norm(s: str) -> str:
    s = (s or "").strip().upper()
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def name_similarity(a: str, b: str) -> float:
    ta = set(norm(a).split())
    tb = set(norm(b).split())
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def canonical_name(name: str) -> str:
    cleaned = COMPANY_SUFFIX_RE.sub(" ", name or "")
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned or (name or "").strip()


def normalize_url(site: str) -> str:
    site = (site or "").strip()
    if not site:
        return ""
    if not site.startswith(("http://", "https://")):
        site = "https://" + site
    return site


def get_domain(site: str) -> str:
    site = normalize_url(site)
    if not site:
        return ""
    try:
        return urlparse(site).netloc.lower().replace("www.", "")
    except Exception:
        return ""


def places_text_search_page(query: str, pagetoken: str = None):
    url = "https://maps.googleapis.com/maps/api/place/textsearch/json"
    params = {"query": query, "key": API_KEY}
    if pagetoken:
        params = {"pagetoken": pagetoken, "key": API_KEY}
    r = SESSION.get(url, params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def places_text_search_all(query: str, max_pages: int = 3):
    results = []
    token = None
    for page_idx in range(max_pages):
        data = places_text_search_page(query, pagetoken=token)
        results.extend(data.get("results", []))
        token = data.get("next_page_token")
        if not token:
            break
        if page_idx < max_pages - 1:
            time.sleep(2.2)
    return results


def places_find_place(query: str):
    url = "https://maps.googleapis.com/maps/api/place/findplacefromtext/json"
    params = {
        "input": query,
        "inputtype": "textquery",
        "fields": "place_id,name,formatted_address,business_status,website,formatted_phone_number,international_phone_number",
        "key": API_KEY,
    }
    r = SESSION.get(url, params=params, timeout=30)
    r.raise_for_status()
    return r.json()


def places_details(place_id: str):
    url = "https://maps.googleapis.com/maps/api/place/details/json"
    fields = (
        "name,website,formatted_phone_number,international_phone_number,"
        "formatted_address,business_status"
    )
    r = SESSION.get(url, params={"place_id": place_id, "fields": fields, "key": API_KEY}, timeout=30)
    r.raise_for_status()
    return r.json()


def score_candidate(target_name: str, domain: str, det: dict) -> float:
    nm = det.get("name", "")
    web = (det.get("website") or "").lower().replace("www.", "")
    status = det.get("business_status", "")

    score = 0.0
    if domain and domain in web:
        score += 10.0
    score += 5.0 * name_similarity(target_name, nm)
    if status == "OPERATIONAL":
        score += 1.0
    if det.get("international_phone_number") or det.get("formatted_phone_number"):
        score += 0.8
    if det.get("formatted_address"):
        score += 0.8
    return score


def build_queries(nome: str, domain: str):
    nome = (nome or "").strip()
    alt = canonical_name(nome)

    queries = []
    if domain:
        queries.append(f"{nome} {domain} Brasil")
        queries.append(f"{alt} {domain} Brasil")
    queries.append(f"{nome} embalagens Brasil")
    queries.append(f"{alt} embalagens Brasil")
    queries.append(f"{nome} papelão ondulado Brasil")
    queries.append(f"{alt} Brasil")
    queries.append(f"{nome} Brasil")

    uniq = []
    seen = set()
    for q in queries:
        q = q.strip()
        if q and q not in seen:
            seen.add(q)
            uniq.append(q)
    return uniq


def copy_via_powershell(src, dst):
    cmd = [
        "powershell",
        "-NoProfile",
        "-Command",
        f"Copy-Item -LiteralPath '{src}' -Destination '{dst}' -Force",
    ]
    subprocess.run(cmd, check=True, capture_output=True, text=True)


def load_workbook_with_lock_fallback(path):
    try:
        wb = openpyxl.load_workbook(path)
        wb_values = openpyxl.load_workbook(path, data_only=True)
        return wb, wb_values, None
    except PermissionError:
        temp_copy = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        temp_copy.close()
        try:
            shutil.copy2(path, temp_copy.name)
        except PermissionError:
            copy_via_powershell(path, temp_copy.name)
        wb = openpyxl.load_workbook(temp_copy.name)
        wb_values = openpyxl.load_workbook(temp_copy.name, data_only=True)
        return wb, wb_values, temp_copy.name


def save_workbook_with_fallback(wb, path):
    try:
        wb.save(path)
        return path
    except PermissionError:
        base, ext = os.path.splitext(path)
        alt = f"{base} - {time.strftime('%Y%m%d_%H%M%S')}{ext or '.xlsx'}"
        wb.save(alt)
        return alt


def fetch_html(url: str) -> str:
    r = SESSION.get(url, timeout=25, allow_redirects=True)
    r.raise_for_status()
    r.encoding = r.apparent_encoding or "utf-8"
    return r.text


def make_soup(html: str):
    try:
        return BeautifulSoup(html, "lxml")
    except FeatureNotFound:
        return BeautifulSoup(html, "html.parser")


def extract_from_jsonld(soup):
    phones = set()
    addresses = set()

    for sc in soup.find_all("script", attrs={"type": "application/ld+json"}):
        raw = (sc.string or "").strip()
        if not raw:
            continue
        try:
            data = json.loads(raw)
        except Exception:
            continue

        items = []
        if isinstance(data, list):
            items = data
        elif isinstance(data, dict) and isinstance(data.get("@graph"), list):
            items = data["@graph"]
        elif isinstance(data, dict):
            items = [data]

        for it in items:
            if not isinstance(it, dict):
                continue

            tel = it.get("telephone") or it.get("phone")
            if isinstance(tel, str):
                for m in PHONE_RE.findall(tel):
                    phones.add(m)

            addr = it.get("address")
            if isinstance(addr, dict):
                parts = [
                    addr.get("streetAddress"),
                    addr.get("addressLocality"),
                    addr.get("addressRegion"),
                    addr.get("postalCode"),
                    addr.get("addressCountry"),
                ]
                built = ", ".join([p for p in parts if p])
                if built:
                    addresses.add(built)
            elif isinstance(addr, str) and addr.strip():
                addresses.add(addr.strip())

    return list(phones), list(addresses)


def extract_phones_and_address_from_text(text: str):
    phones = set(PHONE_RE.findall(text or ""))
    addresses = set()

    for m in ADDR_HINT_RE.finditer(text or ""):
        snippet = m.group(0).strip()
        if len(snippet) >= 15:
            addresses.add(snippet)

    return list(phones), list(addresses)


def best_phone(phones):
    cleaned = []
    for p in phones:
        p = re.sub(r"\s+", " ", (p or "")).strip()
        if p:
            cleaned.append(p)
    if not cleaned:
        return ""
    cleaned = list(dict.fromkeys(cleaned))
    cleaned.sort(key=len, reverse=True)
    return cleaned[0]


def best_address(addresses):
    candidates = [a.strip() for a in addresses if a and a.strip()]
    if not candidates:
        return ""
    candidates.sort(key=len, reverse=True)
    return candidates[0]


def crawl_contact_pages(base_url: str, max_pages: int = 12):
    base_url = normalize_url(base_url)
    if not base_url:
        return []

    common = [
        "/contato",
        "/contato/",
        "/fale-conosco",
        "/fale-conosco/",
        "/contact",
        "/contact/",
        "/contatos",
        "/contatos/",
        "/institucional",
        "/sobre",
        "/unidades",
        "/onde-estamos",
        "/localizacao",
    ]
    candidates = [urljoin(base_url, p) for p in common]

    try:
        soup = make_soup(fetch_html(base_url))
        for a in soup.select("a[href]"):
            href = (a.get("href") or "").strip()
            txt = (a.get_text(" ", strip=True) or "").lower()
            low = href.lower()
            keys = ["contato", "fale", "contact", "unidades", "onde", "local"]
            if any(k in low for k in keys) or any(k in txt for k in keys):
                full = urljoin(base_url, href)
                if urlparse(full).netloc.replace("www.", "") == urlparse(base_url).netloc.replace("www.", ""):
                    candidates.append(full)
    except Exception:
        pass

    uniq = []
    seen = set()
    for u in candidates:
        u = u.split("#")[0]
        if u not in seen:
            seen.add(u)
            uniq.append(u)
        if len(uniq) >= max_pages:
            break

    return uniq


def scrape_site_for_contact(site_url: str):
    site_url = normalize_url(site_url)
    if not site_url:
        return "", "", ""

    pages = crawl_contact_pages(site_url, max_pages=12)
    if site_url not in pages:
        pages.append(site_url)

    best_p = ""
    best_a = ""
    best_src = ""

    for url in pages:
        try:
            soup = make_soup(fetch_html(url))
        except Exception:
            continue

        phones_ld, addrs_ld = extract_from_jsonld(soup)
        text = soup.get_text("\n", strip=True)
        phones_tx, addrs_tx = extract_phones_and_address_from_text(text)

        p = best_phone(phones_ld + phones_tx)
        a = best_address(addrs_ld + addrs_tx)

        score = int(bool(p)) + int(bool(a))
        if score == 2:
            return p, a, url

        if p and not best_p:
            best_p = p
            best_src = url
        if a and not best_a:
            best_a = a
            best_src = url

        time.sleep(0.15)

    return best_p, best_a, best_src


def ensure_col(ws, headers, title):
    key = normalize_header(title)
    if key in headers:
        return headers[key]
    new_col = ws.max_column + 1
    ws.cell(HEADER_ROW, new_col).value = title
    headers[key] = new_col
    return new_col


def normalize_header(text):
    s = str(text or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def build_header_map(ws):
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v:
            headers[normalize_header(v)] = c
    return headers


def parse_number(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def detect_curva_year_columns(ws_curva):
    year_row = None
    year_cols = []
    for r in range(1, min(ws_curva.max_row, 20) + 1):
        cols = []
        for c in range(1, ws_curva.max_column + 1):
            v = ws_curva.cell(r, c).value
            try:
                y = int(str(v).strip())
            except (TypeError, ValueError):
                continue
            if 1900 <= y <= 2100:
                cols.append((c, y))
        if len(cols) >= 2:
            year_row = r
            year_cols = cols
            break
    return year_row, year_cols


def normalize_company_name(name):
    return canonical_name(norm(str(name or "")))


def tokens_name(name):
    parts = [p for p in normalize_company_name(name).split() if len(p) > 1]
    return set(parts)


def build_curva_last_purchase_map(ws_curva):
    _, year_cols = detect_curva_year_columns(ws_curva)
    if not year_cols:
        return {}

    last_purchase = {}
    for r in range(1, ws_curva.max_row + 1):
        raw_name = ws_curva.cell(r, 1).value
        if not raw_name:
            continue

        key = normalize_company_name(raw_name)
        if not key:
            continue

        latest_year = None
        for c, y in year_cols:
            amount = parse_number(ws_curva.cell(r, c).value)
            if amount > 0:
                latest_year = y

        if latest_year:
            prev = last_purchase.get(key)
            if prev is None or latest_year > prev:
                last_purchase[key] = latest_year

    return last_purchase


def similarity_by_tokens(name_a, name_b):
    a = tokens_name(name_a)
    b = tokens_name(name_b)
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def find_best_last_purchase_year(candidate_names, curva_map):
    for n in candidate_names:
        key = normalize_company_name(n)
        if key in curva_map:
            return curva_map[key]

    best_year = None
    best_score = 0.0
    for n in candidate_names:
        for curva_name, year in curva_map.items():
            score = similarity_by_tokens(n, curva_name)
            if score > best_score:
                best_score = score
                best_year = year

    if best_score >= 0.72:
        return best_year
    return None


def fill_base_representantes_last_purchase(wb, wb_values=None):
    if "CURVA ABC" not in wb.sheetnames or "BASE REPRESENTANTES" not in wb.sheetnames:
        return 0

    ws_curva = wb["CURVA ABC"]
    ws_base = wb["BASE REPRESENTANTES"]
    ws_base_values = wb_values["BASE REPRESENTANTES"] if wb_values and "BASE REPRESENTANTES" in wb_values.sheetnames else ws_base

    base_headers = build_header_map(ws_base)

    col_ultima = base_headers.get("ultima compra")
    col_razao = base_headers.get("cliente razao social")
    col_fantasia = base_headers.get("nome fantasia")

    if not col_ultima or not col_razao:
        return 0

    curva_map = build_curva_last_purchase_map(ws_curva)
    if not curva_map:
        return 0

    updated = 0
    for r in range(2, ws_base.max_row + 1):
        razao = ws_base_values.cell(r, col_razao).value
        fantasia = ws_base_values.cell(r, col_fantasia).value if col_fantasia else None

        names = [razao]
        if fantasia:
            names.append(fantasia)

        year = find_best_last_purchase_year(names, curva_map)
        if year:
            ws_base.cell(r, col_ultima).value = str(year)
            updated += 1

    return updated


def build_base_client_name_set(wb, wb_values=None):
    if "BASE REPRESENTANTES" not in wb.sheetnames:
        return set()

    ws_base = wb["BASE REPRESENTANTES"]
    ws_base_values = wb_values["BASE REPRESENTANTES"] if wb_values and "BASE REPRESENTANTES" in wb_values.sheetnames else ws_base
    base_headers = build_header_map(ws_base)
    col_razao = base_headers.get("cliente razao social")
    col_fantasia = base_headers.get("nome fantasia")

    names = set()
    for r in range(2, ws_base.max_row + 1):
        if col_razao:
            v = ws_base_values.cell(r, col_razao).value
            if v:
                names.add(normalize_company_name(v))
        if col_fantasia:
            v = ws_base_values.cell(r, col_fantasia).value
            if v:
                names.add(normalize_company_name(v))
    names.discard("")
    return names


def ensure_removed_headers(ws_removed):
    wanted = ["Tipo da fábrica", "Nome", "Site", "Telefone", "Endereço", "Motivo"]
    current = build_header_map(ws_removed)
    for title in wanted:
        key = normalize_header(title)
        if key not in current:
            col = ws_removed.max_column + 1
            ws_removed.cell(1, col).value = title
            current[key] = col
    return current


def remove_existing_clients_from_clientes(wb, wb_values=None):
    if "Clientes" not in wb.sheetnames or "Removidos" not in wb.sheetnames:
        return 0

    ws_clientes = wb["Clientes"]
    ws_removed = wb["Removidos"]

    client_headers = build_header_map(ws_clientes)

    col_tipo = client_headers.get("tipo da fabrica")
    col_nome = client_headers.get("nome")
    col_site = client_headers.get("site")
    col_tel = client_headers.get("telefone")
    col_end = client_headers.get("endereco")
    if not col_nome:
        return 0

    removed_headers = ensure_removed_headers(ws_removed)
    base_names = build_base_client_name_set(wb, wb_values=wb_values)

    to_remove = []
    for r in range(2, ws_clientes.max_row + 1):
        nome = ws_clientes.cell(r, col_nome).value
        if not nome:
            continue
        key = normalize_company_name(nome)
        if key in base_names:
            to_remove.append(r)
            continue

        # Fuzzy fallback para nomes parecidos.
        match = False
        for base_name in base_names:
            if similarity_by_tokens(nome, base_name) >= 0.78:
                match = True
                break
        if match:
            to_remove.append(r)

    for r in reversed(to_remove):
        new_row = ws_removed.max_row + 1
        ws_removed.cell(new_row, removed_headers["tipo da fabrica"]).value = ws_clientes.cell(r, col_tipo).value if col_tipo else None
        ws_removed.cell(new_row, removed_headers["nome"]).value = ws_clientes.cell(r, col_nome).value
        ws_removed.cell(new_row, removed_headers["site"]).value = ws_clientes.cell(r, col_site).value if col_site else None
        ws_removed.cell(new_row, removed_headers["telefone"]).value = ws_clientes.cell(r, col_tel).value if col_tel else None
        ws_removed.cell(new_row, removed_headers["endereco"]).value = ws_clientes.cell(r, col_end).value if col_end else None
        ws_removed.cell(new_row, removed_headers["motivo"]).value = "Já existe na BASE REPRESENTANTES (CRM)"
        ws_clientes.delete_rows(r, 1)

    return len(to_remove)


has_api = has_api_key()
wb, wb_values, temp_path = load_workbook_with_lock_fallback(ARQ_IN)
ws = wb[SHEET]

updated_last_purchase = fill_base_representantes_last_purchase(wb, wb_values=wb_values)
removed_existing = remove_existing_clients_from_clientes(wb, wb_values=wb_values)

headers = {}
for col in range(1, ws.max_column + 1):
    v = ws.cell(HEADER_ROW, col).value
    if v:
        headers[normalize_header(v)] = col

col_nome = headers.get("nome")
col_site = headers.get("site")
col_tel = headers.get("telefone")
col_end = headers.get("endereco")

col_status = ensure_col(ws, headers, "Status")
col_placeid = ensure_col(ws, headers, "PlaceId")
col_score = ensure_col(ws, headers, "Score")
col_src = ensure_col(ws, headers, "Fonte")

if not col_nome or not col_tel or not col_end:
    raise ValueError("Não achei cabeçalhos 'Nome', 'Telefone' e 'Endereço' na aba Clientes.")

processed = 0
for row in range(HEADER_ROW + 1, ws.max_row + 1):
    nome = ws.cell(row, col_nome).value
    if not nome:
        continue

    if MAX_ROWS > 0 and processed >= MAX_ROWS:
        break

    tel = ws.cell(row, col_tel).value
    end = ws.cell(row, col_end).value
    site = ws.cell(row, col_site).value if col_site else ""
    domain = get_domain(str(site) if site else "")

    if tel and end:
        ws.cell(row, col_status).value = "OK (já preenchido)"
        continue

    best = None  # (score, det, pid, fonte)

    if has_api:
        for q in build_queries(str(nome), domain):
            try:
                results = places_text_search_all(q, max_pages=3)
            except requests.HTTPError as e:
                ws.cell(row, col_status).value = f"ERRO textsearch: {e}"
                continue

            for cand in results[:40]:
                pid = cand.get("place_id")
                if not pid:
                    continue
                try:
                    det = places_details(pid).get("result", {})
                except requests.HTTPError:
                    continue

                score = score_candidate(str(nome), domain, det)
                item = (score, det, pid, "Google Places TextSearch")
                if best is None or item[0] > best[0]:
                    best = item

            if best and best[0] >= 6.0:
                break

            time.sleep(SLEEP)

        if not best:
            try:
                fp = places_find_place(f"{nome} Brasil")
            except requests.HTTPError:
                fp = {}

            for cand in fp.get("candidates", [])[:TOP_N]:
                pid = cand.get("place_id")
                if not pid:
                    continue
                score = score_candidate(str(nome), domain, cand)
                item = (score, cand, pid, "Google Places FindPlace")
                if best is None or item[0] > best[0]:
                    best = item

    if best:
        score, det, pid, fonte = best
        phone = det.get("international_phone_number") or det.get("formatted_phone_number") or ""
        addr = det.get("formatted_address") or ""

        if (not tel) and phone:
            ws.cell(row, col_tel).value = phone
        if (not end) and addr:
            ws.cell(row, col_end).value = addr

        ws.cell(row, col_placeid).value = pid
        ws.cell(row, col_score).value = round(score, 2)
        ws.cell(row, col_src).value = fonte

    tel2 = ws.cell(row, col_tel).value
    end2 = ws.cell(row, col_end).value

    if tel2 and end2:
        ws.cell(row, col_status).value = "OK (Places)" if has_api else "OK (Site)"
        processed += 1
        time.sleep(SLEEP)
        continue

    site_url = str(site).strip() if site else ""
    sp, sa, src_url = scrape_site_for_contact(site_url)

    if (not tel2) and sp:
        ws.cell(row, col_tel).value = sp
    if (not end2) and sa:
        ws.cell(row, col_end).value = sa

    tel3 = ws.cell(row, col_tel).value
    end3 = ws.cell(row, col_end).value

    if tel3 and end3:
        ws.cell(row, col_status).value = "OK (Site)"
        ws.cell(row, col_src).value = src_url or site_url
    elif tel3 or end3:
        ws.cell(row, col_status).value = "PARCIAL (Site)"
        ws.cell(row, col_src).value = src_url or site_url
    else:
        ws.cell(row, col_status).value = "NAO_ENCONTRADO (Places+Site)" if has_api else "NAO_ENCONTRADO (Site)"
        ws.cell(row, col_src).value = src_url or site_url

    processed += 1
    time.sleep(SLEEP)

saved_out = save_workbook_with_fallback(wb, ARQ_OUT)

if temp_path:
    try:
        os.remove(temp_path)
    except OSError:
        pass

try:
    wb_values.close()
except Exception:
    pass

print("Gerado:", saved_out)
print("Modo API:", "ATIVO" if has_api else "DESATIVADO")
print("BASE Ultima compra atualizada:", updated_last_purchase)
print("Clientes movidos para Removidos:", removed_existing)
